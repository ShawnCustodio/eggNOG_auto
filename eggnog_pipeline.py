#!/usr/bin/env python3
"""
eggNOG 7 Annotation Pipeline
------------------------------
Takes a .faa file, runs DIAMOND against eggNOG 7, and produces an annotated .xlsx.

Usage:
    python eggnog_pipeline.py Proteins_Sequence.faa

Output:
    annotated_proteins.xlsx  (written next to the input .faa)

Requires:
    conda install -c bioconda diamond
    pip install pandas openpyxl
"""

import os
import sys
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  PATHS TO YOUR DATABASE FILES  –  edit if needed
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR    = r"C:\Users\SKC180002\Downloads\eggNOG"
DB_DIR      = f"{BASE_DIR}/databases"
REF_FA      = f"{DB_DIR}/e7.proteins.fa"        # eggNOG 7 reference proteins
OG_INFO_TSV = f"{DB_DIR}/e7.og_info_kegg_go.tsv"  # OG → GO + KEGG
FAMILIES_TSV= f"{DB_DIR}/e7.protein_families.tsv"  # protein family → member proteins
WORK_DIR    = f"{BASE_DIR}/results"             # intermediate files land here
DIAMOND_DB  = f"{WORK_DIR}/e7_ref.dmnd"
DIAMOND_OUT = f"{WORK_DIR}/diamond_hits.tsv"
THREADS     = 4
# ─────────────────────────────────────────────────────────────────────────────


# ── STEP 1: DIAMOND ──────────────────────────────────────────────────────────

def run_diamond(query_faa: str):
    os.makedirs(WORK_DIR, exist_ok=True)

    # Build DB once — skipped on subsequent runs
    if not os.path.exists(DIAMOND_DB):
        print("[1/4] Building DIAMOND database (one-time, ~5 min for e7.proteins.fa)...")
        subprocess.run(
            ["diamond", "makedb", "--in", REF_FA, "--db", DIAMOND_DB],
            check=True
        )
    else:
        print("[1/4] DIAMOND database already exists, skipping build.")

    print("      Searching query proteins against eggNOG 7...")
    subprocess.run(
        [
            "diamond", "blastp",
            "--query",           query_faa,
            "--db",              DIAMOND_DB,
            "--out",             DIAMOND_OUT,
            "--outfmt", "6",     "qseqid", "sseqid", "pident", "evalue", "bitscore",
            "--max-target-seqs", "1",        # best hit only
            "--evalue",          "1e-5",
            "--threads",         str(THREADS),
            "--sensitive",
        ],
        check=True
    )
    print(f"      DIAMOND hits written → {DIAMOND_OUT}")


def parse_diamond_hits() -> dict:
    """
    Returns {query_protein_id: best_subject_id}.
    Subject IDs look like '1127695.HMPREF9163_00007' — same format used in the TSVs.
    DIAMOND outputs results sorted by bitscore desc, so the first hit per query is best.
    """
    hits = {}
    with open(DIAMOND_OUT) as fh:
        for line in fh:
            parts = line.strip().split("\t")
            if len(parts) >= 2:
                qid, sid = parts[0], parts[1]
                if qid not in hits:
                    hits[qid] = sid
    print(f"      {len(hits)} query proteins matched to eggNOG 7 hits.")
    return hits


# ── STEP 2: LOAD ANNOTATION TABLES ───────────────────────────────────────────

def load_og_info() -> tuple[dict, dict, dict]:
    """
    e7.og_info_kegg_go.tsv — tab-separated, NO header, columns:
      0  OG@taxlevel|flag     e.g. UNK.N3SL@131567|A-1
      1  OG name              e.g. UNK.N3SL
      2  taxonomic level
      3  # proteins
      4  # species
      5  comma-separated member protein IDs  e.g. 1127695.HMPREF9163_00007,...
      6  KEGG KO terms        e.g. K02586|23.87;K02588|3.48
      7  KEGG KO symbols      e.g. nifD|23.87;nifH|3.48
      8  GO slim terms        e.g. GO:0071941|14.11;GO:0016192|0.35

    Returns three dicts all keyed on protein ID (taxid.proteinID):
        protein_to_og_name  → OG name string
        protein_to_go       → raw GO slim string
        protein_to_kegg     → raw KEGG KO string + symbols
    """
    print("[2/4] Loading e7.og_info_kegg_go.tsv and building lookup index...")
    protein_to_og   = {}
    protein_to_go   = {}
    protein_to_kegg = {}
    protein_to_kegg_sym = {}

    with open(OG_INFO_TSV, encoding="utf-8", errors="replace") as fh:
        for line in fh:
            cols = line.rstrip("\n").split("\t")
            if len(cols) < 6:
                continue

            og_name  = cols[1].strip()
            proteins = cols[5].strip()
            kegg_ko  = cols[6].strip() if len(cols) > 6 else ""
            kegg_sym = cols[7].strip() if len(cols) > 7 else ""
            go_terms = cols[8].strip() if len(cols) > 8 else ""

            for prot in proteins.split(","):
                prot = prot.strip()
                if not prot:
                    continue
                # First hit wins (rows ordered from root → leaves; keep first = broadest)
                if prot not in protein_to_og:
                    protein_to_og[prot]       = og_name
                    protein_to_go[prot]        = go_terms
                    protein_to_kegg[prot]      = kegg_ko
                    protein_to_kegg_sym[prot]  = kegg_sym

    print(f"      Indexed {len(protein_to_og):,} protein → OG mappings.")
    return protein_to_og, protein_to_go, protein_to_kegg, protein_to_kegg_sym


def load_families() -> dict:
    """
    e7.protein_families.tsv — tab-separated, NO header, columns:
      0  protein family name
      1  # proteins
      2  # species
      3  # OGs detected
      4  comma-separated member protein IDs
      5  list of taxids
      6  list of OGs

    Returns:
        protein_to_family  → {taxid.proteinID: family_name}
    """
    print("      Loading e7.protein_families.tsv and building lookup index...")
    protein_to_family = {}

    with open(FAMILIES_TSV, encoding="utf-8", errors="replace") as fh:
        for line in fh:
            cols = line.rstrip("\n").split("\t")
            if len(cols) < 5:
                continue
            family_name = cols[0].strip()
            proteins    = cols[4].strip()

            for prot in proteins.split(","):
                prot = prot.strip()
                if prot and prot not in protein_to_family:
                    protein_to_family[prot] = family_name

    print(f"      Indexed {len(protein_to_family):,} protein → family mappings.")
    return protein_to_family


# ── STEP 3: ANNOTATE ─────────────────────────────────────────────────────────

def parse_terms(raw: str, sep: str = ";") -> str:
    """
    Strips score/weight from 'TERM|score' pairs.
    'GO:0071941|14.11;GO:0016192|0.35'  →  'GO:0071941; GO:0016192'
    'K02586|23.87;K02588|3.48'          →  'K02586; K02588'
    """
    if not raw:
        return ""
    return "; ".join(
        item.split("|")[0].strip()
        for item in raw.split(sep)
        if item.strip()
    )


def read_faa_ids(faa_path: str) -> list[str]:
    """Returns list of protein IDs in the order they appear in the .faa."""
    ids = []
    with open(faa_path) as fh:
        for line in fh:
            if line.startswith(">"):
                ids.append(line[1:].split()[0].strip())
    return ids


def annotate(
    faa_path: str,
    diamond_hits: dict,
    protein_to_og: dict,
    protein_to_go: dict,
    protein_to_kegg: dict,
    protein_to_kegg_sym: dict,
    protein_to_family: dict,
) -> pd.DataFrame:
    print("[3/4] Annotating proteins...")

    protein_ids = read_faa_ids(faa_path)
    rows = []

    for pid in protein_ids:
        hit        = diamond_hits.get(pid, "")          # e.g. 1127695.HMPREF9163_00007
        og_name    = protein_to_og.get(hit, "")
        go_raw     = protein_to_go.get(hit, "")
        kegg_raw   = protein_to_kegg.get(hit, "")
        kegg_sym   = protein_to_kegg_sym.get(hit, "")
        fam_name   = protein_to_family.get(hit, "")

        rows.append({
            "Protein_ID":      pid,
            "Best_Hit":        hit,
            "OG_Name":         og_name,
            "Protein_Family":  fam_name,
            "GO_Terms":        parse_terms(go_raw),
            "KEGG_KO":         parse_terms(kegg_raw),
            "KEGG_Symbols":    parse_terms(kegg_sym),
        })

    df = pd.DataFrame(rows)
    annotated = int((df["OG_Name"] != "").sum())
    print(f"      {annotated} / {len(df)} proteins annotated.")
    return df


# ── STEP 4: WRITE EXCEL ───────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
ALT_FILL  = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
BORDER    = Border(
    left=Side(style="thin",   color="BFBFBF"),
    right=Side(style="thin",  color="BFBFBF"),
    top=Side(style="thin",    color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
COL_WIDTHS = {
    "Protein_ID":     22,
    "Best_Hit":       35,
    "OG_Name":        25,
    "Protein_Family": 28,
    "GO_Terms":       60,
    "KEGG_KO":        25,
    "KEGG_Symbols":   30,
}


def write_excel(df: pd.DataFrame, out_path: str):
    print(f"[4/4] Writing Excel → {out_path}")
    df.to_excel(out_path, index=False, sheet_name="Annotations")

    wb = load_workbook(out_path)
    ws = wb["Annotations"]

    # Style header row
    for cell in ws[1]:
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER

    # Style body rows with alternating row shading
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font      = BODY_FONT
            cell.alignment = LEFT
            cell.border    = BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Column widths
    for i, header in enumerate([ws.cell(1, c).value for c in range(1, ws.max_column + 1)], 1):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(header, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Metric"
    ws2["B1"] = "Value"
    for cell in ws2[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER

    total     = len(df)
    annotated = int((df["OG_Name"]        != "").sum())
    with_go   = int((df["GO_Terms"]        != "").sum())
    with_fam  = int((df["Protein_Family"]  != "").sum())
    with_kegg = int((df["KEGG_KO"]         != "").sum())

    for r, (metric, value) in enumerate([
        ("Total proteins",               total),
        ("Proteins with eggNOG OG hit",  annotated),
        ("Proteins with GO terms",       with_go),
        ("Proteins with protein family", with_fam),
        ("Proteins with KEGG KO",        with_kegg),
        ("Annotation rate (%)",          round(annotated / total * 100, 1) if total else 0),
    ], start=2):
        ws2.cell(r, 1, metric).font = BODY_FONT
        ws2.cell(r, 2, value).font  = BODY_FONT

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 18

    wb.save(out_path)
    print(f"\n✓ Done! Saved to: {out_path}")


# ── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) != 2:
        print("Usage: python eggnog_pipeline.py Proteins_Sequence.faa")
        sys.exit(1)

    faa_path = sys.argv[1]
    if not os.path.exists(faa_path):
        sys.exit(f"ERROR: File not found: {faa_path}")

    out_path = os.path.splitext(faa_path)[0] + "_annotated.xlsx"

    # Step 1: DIAMOND search
    run_diamond(faa_path)
    diamond_hits = parse_diamond_hits()

    # Step 2: Load annotation tables (streamed line-by-line — no RAM overload)
    protein_to_og, protein_to_go, protein_to_kegg, protein_to_kegg_sym = load_og_info()
    protein_to_family = load_families()

    # Step 3: Annotate
    df = annotate(
        faa_path,
        diamond_hits,
        protein_to_og,
        protein_to_go,
        protein_to_kegg,
        protein_to_kegg_sym,
        protein_to_family,
    )

    # Step 4: Write Excel
    write_excel(df, out_path)


if __name__ == "__main__":
    main()